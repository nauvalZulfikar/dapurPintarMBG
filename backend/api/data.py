import io
import json
import os
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func, text

from backend.core.database import (
    engine,
    remote_items,
    remote_trays,
    remote_tray_items,
    remote_scan_errors,
    remote_print_jobs,
    db_list_schools,
)
from backend.services.printing import generate_label, db_create_print_job, create_and_push_job
from backend.services.delivery_optimizer import (
    load_schools_from_json,
    fetch_trays_packed_times,
    assign_trays_to_schools,
)
from backend.utils.auth import get_current_user, get_current_kitchen
from backend.utils.permissions import require_permission
from backend.utils.validators import new_item_id
from backend.utils.datetime_helpers import now_local_iso

router = APIRouter()

PAGE_SIZE = 50

SCHOOLS_FILE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "data", "schools.json",
)


# ── Overview ─────────────────────────────────────────────────────────────────

@router.get("/overview")
async def overview(
    date_filter: Optional[str] = Query(None, alias="date"),
    kitchen: dict = Depends(get_current_kitchen),
):
    today = date.fromisoformat(date_filter) if date_filter else date.today()
    kid = kitchen["id"]
    with engine.connect() as c:
        received = c.execute(
            select(func.count()).select_from(remote_items)
            .where(
                (remote_items.c.created_date_receiving == today) &
                (remote_items.c.kitchen_id == kid)
            )
        ).scalar() or 0
        processed = c.execute(
            select(func.count()).select_from(remote_items)
            .where(
                (remote_items.c.created_date_processing == today) &
                (remote_items.c.kitchen_id == kid)
            )
        ).scalar() or 0
        packed = c.execute(
            select(func.count()).select_from(remote_trays)
            .where(
                (remote_trays.c.created_date_packing == today) &
                (remote_trays.c.kitchen_id == kid)
            )
        ).scalar() or 0
        delivered = c.execute(
            select(func.count()).select_from(remote_trays)
            .where(
                (remote_trays.c.created_date_delivery == today) &
                (remote_trays.c.kitchen_id == kid)
            )
        ).scalar() or 0
    return {
        "items_received": received,
        "items_processed": processed,
        "trays_packed": packed,
        "trays_delivered": delivered,
        "date": str(today),
        "kitchen": {"id": kitchen["id"], "slug": kitchen["slug"], "name": kitchen["name"]},
    }


# ── Items ────────────────────────────────────────────────────────────────────

@router.get("/items")
async def get_items(
    date_filter: Optional[str] = Query(None, alias="date"),
    page: int = Query(1, ge=1),
    search: Optional[str] = Query(None),
    include_availability: bool = Query(False),
    kitchen: dict = Depends(get_current_kitchen),
):
    offset = (page - 1) * PAGE_SIZE
    kid = kitchen["id"]
    q = select(remote_items).where(remote_items.c.kitchen_id == kid).order_by(remote_items.c.created_at_receiving.desc())
    count_q = select(func.count()).select_from(remote_items).where(remote_items.c.kitchen_id == kid)

    if date_filter:
        q = q.where(remote_items.c.created_date_receiving == date_filter)
        count_q = count_q.where(remote_items.c.created_date_receiving == date_filter)
    if search:
        q = q.where(remote_items.c.name.ilike(f"%{search}%"))
        count_q = count_q.where(remote_items.c.name.ilike(f"%{search}%"))

    with engine.connect() as c:
        total = c.execute(count_q).scalar() or 0
        rows = c.execute(q.limit(PAGE_SIZE).offset(offset)).fetchall()

    defect_map: dict = {}
    if include_availability and rows:
        ids = [r.id for r in rows]
        with engine.connect() as c:
            sums = c.execute(text("""
                SELECT item_id, COALESCE(SUM(weight_grams), 0) AS total
                FROM defect_items
                WHERE item_id = ANY(:ids)
                GROUP BY item_id
            """), {"ids": ids}).fetchall()
            defect_map = {r.item_id: int(r.total or 0) for r in sums}

    items = []
    for r in rows:
        row = {
            "id": r.id,
            "name": r.name,
            "weight_grams": r.weight_grams,
            "unit": r.unit,
            "reason": r.reason,
            "receiving": r.receiving,
            "created_at_receiving": str(r.created_at_receiving) if r.created_at_receiving else None,
            "created_date_receiving": str(r.created_date_receiving) if r.created_date_receiving else None,
            "processing": r.processing,
            "created_at_processing": str(r.created_at_processing) if r.created_at_processing else None,
        }
        if include_availability:
            already = defect_map.get(r.id, 0)
            row["already_defected_grams"] = already
            row["available_grams"] = max(0, int(r.weight_grams or 0) - already)
        items.append(row)

    return {
        "items": items,
        "page": page,
        "page_size": PAGE_SIZE,
        "total": total,
        "total_pages": max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE),
    }


class CreateItemRequest(BaseModel):
    name: str
    weight: float
    unit: str = "g"
    checklist: Optional[dict] = None
    notes: Optional[str] = None


class UpdateItemRequest(BaseModel):
    name: Optional[str] = None
    weight: Optional[float] = None
    unit: Optional[str] = None


def _print_then_save(item_id, name, weight_g, unit, reason, label, kitchen_id, printer_name):
    """Background: print first, then save to DB."""
    import threading, logging
    log = logging.getLogger(__name__)

    from backend.services.printing import _send_raw_to_printer, LOCAL_PRINT, HAS_WIN32
    if LOCAL_PRINT and HAS_WIN32:
        try:
            _send_raw_to_printer(label, printer_name=printer_name)
        except Exception as e:
            log.error(f"[PRINT] Direct print failed: {e}")

    def _save():
        try:
            with engine.begin() as c:
                c.execute(remote_items.insert().values(
                    id=item_id,
                    kitchen_id=kitchen_id,
                    name=name,
                    weight_grams=weight_g,
                    unit=unit,
                    reason=reason,
                    receiving=True,
                    created_at_receiving=datetime.now(),
                    created_date_receiving=date.today(),
                ))
        except Exception as e:
            log.error(f"[ITEM] DB save failed for {item_id}: {e}")

    threading.Thread(target=_save, daemon=True).start()


@router.post("/items")
async def create_item(
    body: CreateItemRequest,
    background_tasks: BackgroundTasks,
    kitchen: dict = Depends(require_permission("items.create")),
):
    item_id = new_item_id()

    weight_g = int(body.weight)
    if body.unit == "kg":
        weight_g = int(body.weight * 1000)

    reason_data = {}
    if body.checklist:
        reason_data["checklist"] = body.checklist
    if body.notes:
        reason_data["notes"] = body.notes
    reason = json.dumps(reason_data) if reason_data else None

    label = generate_label(item_id, body.name, weight_g, kitchen=kitchen)

    background_tasks.add_task(
        _print_then_save, item_id, body.name, weight_g, body.unit, reason, label,
        kitchen["id"], kitchen.get("printer_name"),
    )

    return {
        "id": item_id,
        "name": body.name,
        "weight_grams": weight_g,
        "unit": body.unit,
    }


@router.post("/items/test-print")
async def test_print_item(kitchen: dict = Depends(require_permission("items.create"))):
    """Print a sample label to verify printer connectivity. Does NOT touch the
    database (no items row, no print_jobs row). The label uses a TEST-* id so
    it can never be confused with a real ingredient."""
    import secrets as _secrets
    from backend.services.printing import LOCAL_PRINT, HAS_WIN32, _send_raw_to_printer
    from backend.api.print_queue import push_job_to_agent

    test_id = "TEST-" + _secrets.token_hex(4).upper()
    label = generate_label(test_id, "TEST PRINT", 0, kitchen=kitchen)

    if LOCAL_PRINT and HAS_WIN32:
        try:
            _send_raw_to_printer(label, printer_name=kitchen.get("printer_name"))
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Printer error: {e}")
        return {"ok": True, "test_id": test_id, "mode": "local"}

    # Cloud-print mode: push directly via WS without queueing in print_jobs.
    # job_id=0 is a sentinel: the WS ack handler ignores falsy ids so no DB row
    # gets marked.
    pushed = await push_job_to_agent(0, label, kitchen_id=kitchen["id"])
    if not pushed:
        raise HTTPException(
            status_code=503,
            detail="Printer agent offline. Check the scanner/printer connection.",
        )
    return {"ok": True, "test_id": test_id, "mode": "ws"}


@router.put("/items/{item_id}")
async def update_item(
    item_id: str,
    body: UpdateItemRequest,
    kitchen: dict = Depends(require_permission("items.edit")),
):
    values = {}
    if body.name is not None:
        values["name"] = body.name
    if body.weight is not None:
        weight_g = int(body.weight)
        if body.unit == "kg":
            weight_g = int(body.weight * 1000)
        values["weight_grams"] = weight_g
    if body.unit is not None:
        values["unit"] = body.unit

    if not values:
        raise HTTPException(status_code=400, detail="No fields to update")

    with engine.begin() as c:
        result = c.execute(
            remote_items.update()
            .where(
                (remote_items.c.id == item_id) &
                (remote_items.c.kitchen_id == kitchen["id"])
            )
            .values(**values)
        )
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Item not found")

    return {"ok": True}


@router.delete("/items/{item_id}")
async def delete_item(item_id: str, kitchen: dict = Depends(require_permission("items.edit"))):
    with engine.begin() as c:
        result = c.execute(
            remote_items.delete()
            .where(
                (remote_items.c.id == item_id) &
                (remote_items.c.kitchen_id == kitchen["id"])
            )
        )
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Item not found")

    return {"ok": True}


# ── Trays ────────────────────────────────────────────────────────────────────

@router.get("/trays")
async def get_trays(
    date_filter: Optional[str] = Query(None, alias="date"),
    page: int = Query(1, ge=1),
    kitchen: dict = Depends(get_current_kitchen),
):
    offset = (page - 1) * PAGE_SIZE
    kid = kitchen["id"]
    q = select(remote_trays).where(remote_trays.c.kitchen_id == kid).order_by(remote_trays.c.created_at_packing.desc().nullslast())
    count_q = select(func.count()).select_from(remote_trays).where(remote_trays.c.kitchen_id == kid)

    if date_filter:
        q = q.where(remote_trays.c.created_date_packing == date_filter)
        count_q = count_q.where(remote_trays.c.created_date_packing == date_filter)

    with engine.connect() as c:
        total = c.execute(count_q).scalar() or 0
        rows = c.execute(q.limit(PAGE_SIZE).offset(offset)).fetchall()

    trays = []
    for r in rows:
        trays.append({
            "id": r.id,
            "tray_id": r.tray_id,
            "reason": r.reason,
            "packing": r.packing,
            "created_at_packing": str(r.created_at_packing) if r.created_at_packing else None,
            "delivery": r.delivery,
            "created_at_delivery": str(r.created_at_delivery) if r.created_at_delivery else None,
        })

    return {
        "trays": trays,
        "page": page,
        "page_size": PAGE_SIZE,
        "total": total,
        "total_pages": max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE),
    }


# ── Delivery ─────────────────────────────────────────────────────────────────

MEALS_PER_SCAN = 10


def _compute_deliveries(n_scans: int, schools_sorted: list, delivery_rows: list) -> list:
    meals: dict = {s["school_id"]: 0 for s in schools_sorted}
    last_idx: dict = {}

    remaining = n_scans
    scan_cursor = 0
    for school in schools_sorted:
        sid = school["school_id"]
        n_full = int(school["student_count"]) // MEALS_PER_SCAN
        used = min(n_full, remaining)
        if used > 0:
            meals[sid] += used * MEALS_PER_SCAN
            last_idx[sid] = scan_cursor + used - 1
        scan_cursor += n_full
        remaining -= used
        if remaining <= 0:
            break

    total_full_scans = sum(int(s["student_count"]) // MEALS_PER_SCAN for s in schools_sorted)

    if remaining > 0:
        combined_avail = remaining * MEALS_PER_SCAN
        cum_rem = 0
        for school in schools_sorted:
            sid = school["school_id"]
            remainder = int(school["student_count"]) % MEALS_PER_SCAN
            if remainder == 0:
                continue
            take = min(remainder, combined_avail)
            if take > 0:
                meals[sid] += take
                last_meal_pos = cum_rem + take
                last_combined_scan = (last_meal_pos - 1) // MEALS_PER_SCAN
                last_idx[sid] = total_full_scans + last_combined_scan
            combined_avail -= take
            cum_rem += remainder
            if combined_avail <= 0:
                break

    result = []
    for school in schools_sorted:
        sid = school["school_id"]
        last_delivered_at = None
        idx = last_idx.get(sid)
        if idx is not None and idx < len(delivery_rows):
            ts = delivery_rows[idx].created_at_delivery
            last_delivered_at = str(ts) if ts else None
        result.append({
            "school_id": int(sid),
            "school_name": school["name"],
            "student_count": int(school["student_count"]),
            "meals_delivered": meals[sid],
            "distance": school["distance"],
            "last_delivered_at": last_delivered_at,
        })
    return result


@router.get("/delivery")
async def get_delivery(
    date_filter: Optional[str] = Query(None, alias="date"),
    kitchen: dict = Depends(get_current_kitchen),
):
    target_date = date.fromisoformat(date_filter) if date_filter else date.today()

    # Phase 1: schools come from DB (kitchen-scoped). JSON file is now seed-only.
    schools_raw = db_list_schools(kitchen["id"], active_only=True)
    schools_sorted = sorted(schools_raw, key=lambda s: s["distance"])

    with engine.connect() as c:
        delivery_rows = c.execute(text("""
            SELECT tray_id, created_at_delivery FROM trays
            WHERE created_date_delivery = :d AND delivery = true AND kitchen_id = :kid
            ORDER BY created_at_delivery ASC
        """), {"d": str(target_date), "kid": kitchen["id"]}).fetchall()

    result = _compute_deliveries(len(delivery_rows), schools_sorted, delivery_rows)
    return {"assignments": result, "date": str(target_date)}


# ── Stats ─────────────────────────────────────────────────────────────────────

@router.get("/stats")
async def get_stats(
    date_filter: Optional[str] = Query(None, alias="date"),
    kitchen: dict = Depends(get_current_kitchen),
):
    from datetime import timedelta
    target = date.fromisoformat(date_filter) if date_filter else date.today()
    week_start = target - timedelta(days=6)
    kid = kitchen["id"]

    with engine.connect() as c:
        hourly_rows = c.execute(text("""
            SELECT hour, SUM(cnt) AS scans FROM (
                SELECT EXTRACT(HOUR FROM created_at_receiving)::int AS hour, COUNT(*) AS cnt
                FROM items WHERE created_date_receiving = :d AND created_at_receiving IS NOT NULL AND kitchen_id = :kid GROUP BY 1
                UNION ALL
                SELECT EXTRACT(HOUR FROM created_at_processing)::int, COUNT(*)
                FROM items WHERE created_date_processing = :d AND created_at_processing IS NOT NULL AND kitchen_id = :kid GROUP BY 1
                UNION ALL
                SELECT EXTRACT(HOUR FROM created_at_packing)::int, COUNT(*)
                FROM trays WHERE created_date_packing = :d AND created_at_packing IS NOT NULL AND kitchen_id = :kid GROUP BY 1
                UNION ALL
                SELECT EXTRACT(HOUR FROM created_at_delivery)::int, COUNT(*)
                FROM trays WHERE created_date_delivery = :d AND created_at_delivery IS NOT NULL AND kitchen_id = :kid GROUP BY 1
            ) sub GROUP BY hour ORDER BY hour
        """), {"d": str(target), "kid": kid}).fetchall()

        trend_rows = c.execute(text("""
            SELECT created_date_receiving AS day, COUNT(*) AS received
            FROM items
            WHERE created_date_receiving >= :start AND created_date_receiving <= :end AND kitchen_id = :kid
            GROUP BY created_date_receiving ORDER BY created_date_receiving
        """), {"start": str(week_start), "end": str(target), "kid": kid}).fetchall()

        r2p = c.execute(text("""
            SELECT AVG(EXTRACT(EPOCH FROM (created_at_processing - created_at_receiving)) / 60)
            FROM items
            WHERE created_date_receiving = :d AND kitchen_id = :kid
              AND created_at_processing IS NOT NULL AND created_at_receiving IS NOT NULL
        """), {"d": str(target), "kid": kid}).scalar()

        p2d = c.execute(text("""
            SELECT AVG(EXTRACT(EPOCH FROM (created_at_delivery - created_at_packing)) / 60)
            FROM trays
            WHERE created_date_packing = :d AND kitchen_id = :kid
              AND created_at_delivery IS NOT NULL AND created_at_packing IS NOT NULL
        """), {"d": str(target), "kid": kid}).scalar()

    schools = load_schools_from_json(SCHOOLS_FILE)
    total_students = sum(int(s.student_count) for s in schools)

    return {
        "date": str(target),
        "hourly_scans": [{"hour": int(r[0]), "scans": int(r[1])} for r in hourly_rows],
        "weekly_trend": [{"date": str(r[0]), "received": int(r[1])} for r in trend_rows],
        "avg_receive_to_process_mins": round(float(r2p), 1) if r2p else None,
        "avg_pack_to_deliver_mins": round(float(p2d), 1) if p2d else None,
        "total_students": total_students,
    }


# ── Scan Errors ──────────────────────────────────────────────────────────────

@router.get("/scan-errors")
async def get_scan_errors(
    page: int = Query(1, ge=1),
    kitchen: dict = Depends(require_permission("scan_errors.view")),
):
    offset = (page - 1) * PAGE_SIZE
    kid = kitchen["id"]
    with engine.connect() as c:
        total = c.execute(
            select(func.count()).select_from(remote_scan_errors)
            .where(remote_scan_errors.c.kitchen_id == kid)
        ).scalar() or 0
        rows = c.execute(
            select(remote_scan_errors)
            .where(remote_scan_errors.c.kitchen_id == kid)
            .order_by(remote_scan_errors.c.id.desc())
            .limit(PAGE_SIZE).offset(offset)
        ).fetchall()

    errors = []
    for r in rows:
        errors.append({
            "id": r.id,
            "code": r.code,
            "step": r.step,
            "created_at": r.created_at,
            "reason": r.reason,
        })

    return {
        "errors": errors,
        "page": page,
        "page_size": PAGE_SIZE,
        "total": total,
        "total_pages": max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE),
    }


# ── Schools ──────────────────────────────────────────────────────────────────

@router.get("/schools")
async def get_schools(
    user: dict = Depends(get_current_user),
    kitchen: dict = Depends(get_current_kitchen),
):
    """Phase 1: kitchen-scoped DB lookup. Returns same shape as legacy JSON,
    so existing frontend callers (Dashboard, NutritionReport, Countdown) keep
    working. Inactive schools are filtered out for the operational endpoint;
    use `/api/admin/schools?include_inactive=true` for the admin list.
    """
    return {"schools": db_list_schools(kitchen["id"], active_only=True)}


# ── Countdown (public — no auth) ──────────────────────────────────────────────

SAFE_HOURS = 4


@router.get("/countdown/{tray_id}")
async def get_countdown(tray_id: str):
    """Public endpoint: since tray_id is not globally unique across kitchens,
    we resolve the most recently delivered tray with that id across all kitchens."""
    from datetime import timedelta
    from backend.api.scans import _scan_allocations

    with engine.connect() as c:
        row = c.execute(
            select(
                remote_trays.c.delivery,
                remote_trays.c.created_at_delivery,
                remote_trays.c.created_date_delivery,
                remote_trays.c.kitchen_id,
            )
            .where(remote_trays.c.tray_id == tray_id)
            .order_by(remote_trays.c.created_at_delivery.desc().nullslast())
            .limit(1)
        ).first()

    if row is None:
        raise HTTPException(status_code=404, detail="Tray not found")

    delivered_at = None
    safe_until = None
    allocations = []

    if row.delivery and row.created_at_delivery:
        from backend.core.config import TZ_REGION
        try:
            from zoneinfo import ZoneInfo
            tz = ZoneInfo(TZ_REGION)
        except Exception:
            tz = None
        delivered_at_dt = row.created_at_delivery
        if tz and delivered_at_dt.tzinfo is None:
            delivered_at_dt = delivered_at_dt.replace(tzinfo=tz)
        safe_until_dt = delivered_at_dt + timedelta(hours=SAFE_HOURS)
        delivered_at = delivered_at_dt.isoformat()
        safe_until = safe_until_dt.isoformat()

        with engine.connect() as c:
            scan_rows = c.execute(text("""
                SELECT tray_id FROM trays
                WHERE created_date_delivery = :d AND delivery = true AND kitchen_id = :kid
                ORDER BY created_at_delivery ASC
            """), {"d": str(row.created_date_delivery), "kid": row.kitchen_id}).fetchall()

        scan_order = [r[0] for r in scan_rows]
        try:
            n = scan_order.index(tray_id) + 1
        except ValueError:
            n = len(scan_order)

        with open(SCHOOLS_FILE, "r", encoding="utf-8") as f:
            schools = json.load(f)
        schools_sorted = sorted(schools, key=lambda s: s["distance"])
        allocations = _scan_allocations(n, schools_sorted)

    return {
        "tray_id": tray_id,
        "delivered_at": delivered_at,
        "safe_until": safe_until,
        "allocations": allocations,
    }


# ── Export Laporan Harian (Excel) ─────────────────────────────────────────────

@router.get("/export/daily")
async def export_daily(
    date_filter: Optional[str] = Query(None, alias="date"),
    kitchen: dict = Depends(require_permission("export.daily")),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    target = date.fromisoformat(date_filter) if date_filter else date.today()
    kid = kitchen["id"]

    with engine.connect() as c:
        items_rows = c.execute(text("""
            SELECT id, name, weight_grams, unit,
                   created_at_receiving, created_at_processing
            FROM items
            WHERE created_date_receiving = :d AND kitchen_id = :kid
            ORDER BY created_at_receiving ASC
        """), {"d": str(target), "kid": kid}).fetchall()

        tray_rows = c.execute(text("""
            SELECT tray_id, created_at_packing, created_at_delivery
            FROM trays
            WHERE (created_date_packing = :d OR created_date_delivery = :d)
              AND kitchen_id = :kid
            ORDER BY created_at_packing ASC
        """), {"d": str(target), "kid": kid}).fetchall()

    wb = openpyxl.Workbook()

    ws_sum = wb.active
    ws_sum.title = "Ringkasan"
    header_fill = PatternFill("solid", fgColor="1B3A6B")
    header_font = Font(bold=True, color="FFFFFF")

    ws_sum.append([f"Laporan Harian — {kitchen['name']}", str(target)])
    ws_sum.append([])
    ws_sum.append(["Metrik", "Jumlah"])
    for cell in ws_sum[3]:
        cell.font = header_font
        cell.fill = header_fill

    received  = sum(1 for r in items_rows)
    processed = sum(1 for r in items_rows if r[5] is not None)
    packed    = sum(1 for r in tray_rows if r[1] is not None)
    delivered = sum(1 for r in tray_rows if r[2] is not None)

    ws_sum.append(["Item Diterima",   received])
    ws_sum.append(["Item Diproses",   processed])
    ws_sum.append(["Tray Dipacking",  packed])
    ws_sum.append(["Tray Dikirim",    delivered])
    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 15

    ws_items = wb.create_sheet("Item Bahan")
    headers = ["ID", "Nama", "Berat (g)", "Satuan", "Waktu Terima", "Waktu Proses"]
    ws_items.append(headers)
    for cell in ws_items[1]:
        cell.font = header_font
        cell.fill = header_fill
    for r in items_rows:
        ws_items.append([r[0], r[1], r[2], r[3],
                         str(r[4]) if r[4] else "", str(r[5]) if r[5] else ""])
    for col, width in zip(["A","B","C","D","E","F"], [16, 25, 10, 8, 22, 22]):
        ws_items.column_dimensions[col].width = width

    ws_trays = wb.create_sheet("Tray")
    headers = ["Tray ID", "Waktu Packing", "Waktu Delivery"]
    ws_trays.append(headers)
    for cell in ws_trays[1]:
        cell.font = header_font
        cell.fill = header_fill
    for r in tray_rows:
        ws_trays.append([r[0], str(r[1]) if r[1] else "", str(r[2]) if r[2] else ""])
    for col, width in zip(["A","B","C"], [16, 22, 22]):
        ws_trays.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"laporan_{kitchen['slug']}_{target}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Export laporan rentang tanggal (weekly / monthly / custom) ────────────────

@router.get("/export/range")
async def export_range(
    from_date: str = Query(..., alias="from"),
    to_date: str = Query(..., alias="to"),
    kitchen: dict = Depends(require_permission("export.range")),
):
    """Excel report across a date range with per-day breakdown.

    Use for weekly/monthly reconciliation. Accountant & admin only.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    try:
        d_from = date.fromisoformat(from_date)
        d_to = date.fromisoformat(to_date)
    except ValueError:
        raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD.")
    if d_from > d_to:
        raise HTTPException(400, "'from' must be <= 'to'")
    span_days = (d_to - d_from).days + 1
    if span_days > 366:
        raise HTTPException(400, "Range cannot exceed 366 days")

    kid = kitchen["id"]
    with engine.connect() as c:
        per_day = c.execute(text("""
            SELECT CAST(d AS date) AS day,
                   COALESCE(i.received,  0)  AS received,
                   COALESCE(i.processed, 0)  AS processed,
                   COALESCE(t.packed,    0)  AS packed,
                   COALESCE(t.delivered, 0)  AS delivered
            FROM generate_series(CAST(:f AS date), CAST(:t AS date), interval '1 day') AS d
            LEFT JOIN (
                SELECT created_date_receiving AS day,
                       COUNT(*) AS received,
                       COUNT(created_at_processing) AS processed
                FROM items
                WHERE kitchen_id = :kid
                  AND created_date_receiving BETWEEN CAST(:f AS date) AND CAST(:t AS date)
                GROUP BY created_date_receiving
            ) i ON i.day = CAST(d AS date)
            LEFT JOIN (
                SELECT created_date_packing AS day,
                       COUNT(*) AS packed,
                       COUNT(created_at_delivery) AS delivered
                FROM trays
                WHERE kitchen_id = :kid
                  AND created_date_packing BETWEEN CAST(:f AS date) AND CAST(:t AS date)
                GROUP BY created_date_packing
            ) t ON t.day = CAST(d AS date)
            ORDER BY d
        """), {"f": str(d_from), "t": str(d_to), "kid": kid}).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ringkasan"
    header_fill = PatternFill("solid", fgColor="1B3A6B")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append([f"Laporan {d_from} s/d {d_to} — {kitchen['name']}"])
    ws.append([])
    ws.append(["Tanggal", "Diterima", "Diproses", "Packed", "Dikirim"])
    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill

    tot = [0, 0, 0, 0]
    for r in per_day:
        ws.append([str(r.day), r.received, r.processed, r.packed, r.delivered])
        tot[0] += r.received; tot[1] += r.processed
        tot[2] += r.packed;   tot[3] += r.delivered

    ws.append([])
    ws.append(["TOTAL", *tot])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col, w in zip(["A","B","C","D","E"], [14, 12, 12, 12, 12]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"laporan_{kitchen['slug']}_{d_from}_to_{d_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Waste / variance report (accountant / admin) ─────────────────────────────

@router.get("/reports/variance")
async def variance_report(
    from_date: str = Query(..., alias="from"),
    to_date: str = Query(..., alias="to"),
    kitchen: dict = Depends(require_permission("reports.variance")),
):
    """Variance & waste summary for a date range.

    Returns per-day counts + aggregates:
      items_received → items_processed → trays_packed → trays_delivered
    plus variance gaps (processed/received, delivered/packed) as percent.
    """
    try:
        d_from = date.fromisoformat(from_date)
        d_to = date.fromisoformat(to_date)
    except ValueError:
        raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD.")
    if d_from > d_to:
        raise HTTPException(400, "'from' must be <= 'to'")
    if (d_to - d_from).days > 366:
        raise HTTPException(400, "Range cannot exceed 366 days")

    kid = kitchen["id"]
    with engine.connect() as c:
        per_day = c.execute(text("""
            SELECT CAST(d AS date) AS day,
                   COALESCE(i.received,  0)  AS received,
                   COALESCE(i.processed, 0)  AS processed,
                   COALESCE(t.packed,    0)  AS packed,
                   COALESCE(t.delivered, 0)  AS delivered
            FROM generate_series(CAST(:f AS date), CAST(:t AS date), interval '1 day') AS d
            LEFT JOIN (
                SELECT created_date_receiving AS day,
                       COUNT(*) AS received,
                       COUNT(created_at_processing) AS processed
                FROM items
                WHERE kitchen_id = :kid
                  AND created_date_receiving BETWEEN CAST(:f AS date) AND CAST(:t AS date)
                GROUP BY created_date_receiving
            ) i ON i.day = CAST(d AS date)
            LEFT JOIN (
                SELECT created_date_packing AS day,
                       COUNT(*) AS packed,
                       COUNT(created_at_delivery) AS delivered
                FROM trays
                WHERE kitchen_id = :kid
                  AND created_date_packing BETWEEN CAST(:f AS date) AND CAST(:t AS date)
                GROUP BY created_date_packing
            ) t ON t.day = CAST(d AS date)
            ORDER BY d
        """), {"f": str(d_from), "t": str(d_to), "kid": kid}).fetchall()

    days = []
    tot = {"received": 0, "processed": 0, "packed": 0, "delivered": 0}
    for r in per_day:
        rec, pro, pkd, dlv = r.received, r.processed, r.packed, r.delivered
        days.append({
            "date": str(r.day),
            "received":  rec,
            "processed": pro,
            "packed":    pkd,
            "delivered": dlv,
            "processed_pct": round(100.0 * pro / rec, 1) if rec else None,
            "delivered_pct": round(100.0 * dlv / pkd, 1) if pkd else None,
        })
        tot["received"] += rec
        tot["processed"] += pro
        tot["packed"] += pkd
        tot["delivered"] += dlv

    def _pct(num, den):
        return round(100.0 * num / den, 1) if den else None

    summary = {
        **tot,
        "processed_pct":  _pct(tot["processed"], tot["received"]),
        "delivered_pct":  _pct(tot["delivered"], tot["packed"]),
        "processing_waste_pct": _pct(tot["received"] - tot["processed"], tot["received"]),
        "delivery_waste_pct":   _pct(tot["packed"] - tot["delivered"], tot["packed"]),
    }
    return {
        "kitchen": {"id": kitchen["id"], "slug": kitchen["slug"], "name": kitchen["name"]},
        "from": str(d_from),
        "to": str(d_to),
        "days": days,
        "summary": summary,
    }
