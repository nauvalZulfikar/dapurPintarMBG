"""Generate docs/DPMBG_Features.xlsx — feature catalog grouped by role.

Path column uses **actual UI labels** from the sidebar / page tabs (mirrors the
running app at http://localhost:5173/), not URL slugs. Each row reads like a
breadcrumb: "Sidebar > Page > Tab > Action".

Run:
    python -m backend.scripts.build_features_xlsx
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ── Style helpers ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")          # navy
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
GROUP_FILL  = PatternFill("solid", fgColor="FEF3C7")          # amber-100
GROUP_FONT  = Font(bold=True, color="92400E", size=11)
ALT_FILL    = PatternFill("solid", fgColor="F8FAFC")          # slate-50
THIN        = Side(border_style="thin", color="CBD5E1")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP        = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER      = Alignment(wrap_text=True, vertical="center", horizontal="center")


# ── Column schema ──────────────────────────────────────────────────────────
COLS = [
    ("No",                 6),
    ("Fitur",              26),
    ("Path UI",            48),   # "Sidebar > Page > Tab > Button"
    ("URL",                26),
    ("Cara Pakai (langkah detail)", 70),
    ("Hasil yang Diharapkan",       55),
    ("Endpoint API",                36),
    ("Test Data Sandbox",           34),
    ("Catatan",                     32),
]


# ── Feature catalog ────────────────────────────────────────────────────────
# Each role group is a list of feature dicts. Keep cara/hasil concrete and step-
# by-step. Path uses the literal labels visible in the sidebar after login.
ROLE_GROUPS: list[tuple[str, list[dict]]] = [
    # ═════════════════════════════════════════════════════════════════════
    ("🌐 SEMUA ROLE — fitur yang muncul untuk semua user yang login", [
        {
            "fitur": "Login",
            "path": "Login Page (sebelum sidebar muncul)",
            "url": "/login",
            "cara": (
                "1. Buka http://localhost:5173/login\n"
                "2. Form: input field 'Username' → ketik 'testadmin'\n"
                "3. Field 'Password' → ketik 'testadmin123'\n"
                "4. (Opsional) Field 'Org slug' → 'testsandbox' kalau username ada di >1 org\n"
                "5. Klik tombol biru 'Sign In'\n"
                "6. Loading spinner ~15-20 detik (login pertama setelah backend idle, audit log ke Supabase)"
            ),
            "hasil": (
                "• Auto-redirect ke '/' (Dashboard utama)\n"
                "• Sidebar gelap navy muncul di kiri (lebar 240px)\n"
                "• Header brand: 'DAPUR PINTAR' + nama kitchen\n"
                "• localStorage tersimpan: token JWT (exp 7 hari), user, kitchens, permissions\n"
                "• Rate limit: 10 attempt / menit per IP — kalo lebih, 429"
            ),
            "api": "POST /api/auth/login",
            "data": "testadmin / testadmin123 / testsandbox",
            "note": "Login pertama lambat = audit log INSERT ke remote Supabase pooler. Selanjutnya cached.",
        },
        {
            "fitur": "Notifikasi Bell",
            "path": "Header Sidebar > 🔔 (icon kanan-atas, dekat tombol collapse)",
            "url": "(seluruh halaman, persistent)",
            "cara": (
                "1. Klik icon 🔔 di pojok kanan-atas sidebar (di samping tombol collapse '←')\n"
                "2. Dropdown panel muncul (lebar ~360px, list scrollable)\n"
                "3. Setiap notif punya: emoji type + title + timestamp + link target\n"
                "4. Klik notif → auto-mark-as-read + navigate ke link\n"
                "5. Klik 'Mark all read' (di footer dropdown) untuk clear semua\n"
                "6. Settings → ⚙️ subscription (event types: menu.pending_review, inspection.scheduled, batch.qc_ready, dst.)"
            ),
            "hasil": (
                "• Real-time delivery via Server-Sent Events (SSE) — connection ke /api/sse\n"
                "• Badge angka merah di icon bell kalau ada unread\n"
                "• Heartbeat tiap 30s (keep-alive)\n"
                "• Notif disimpan max 30 hari di tabel notifications\n"
                "• Toast popup 5 detik untuk severity high"
            ),
            "api": "GET /api/notifications, POST /api/notifications/{id}/read, POST /mark-all-read, /subscriptions",
            "data": "Notif auto-generate setiap menu submitted / inspection / batch QC / dispute",
            "note": "PWA service-worker ngeblok kalau dev mode — refresh hard (Ctrl+Shift+R) kalau stuck.",
        },
        {
            "fitur": "Dashboard Utama",
            "path": "Sidebar > Dashboard (⊞)",
            "url": "/",
            "cara": (
                "1. Klik 'Dashboard' di sidebar paling atas (icon ⊞)\n"
                "2. Halaman menampilkan grid metric cards:\n"
                "   - 'Target Porsi Hari Ini' (angka biru besar)\n"
                "   - 'Items Diterima' (warna biru)\n"
                "   - 'Tray Delivered'\n"
                "   - 'Confirmed Guru' (warna hijau)\n"
                "   - 'Defect Rate %' (merah kalau >5%, hijau kalau ≤5%)\n"
                "   - 'Expense Hari Ini' (Rp format ID)\n"
                "3. Section di bawah: shortcut tombol ke Approval Menu / Inspection / Production"
            ),
            "hasil": (
                "• KPI snapshot real-time hari ini\n"
                "• Card warna-coded: red >5% defect rate; green confirmed guru\n"
                "• Chart bar/line trend 7 hari terakhir (kalau data cukup)\n"
                "• Compare button → /executive untuk drill-down"
            ),
            "api": "GET /api/executive/kpi",
            "data": "Hari ini: 5 porsi target (dari batch sandbox), 4 leftover, 2 expenses",
            "note": "Refresh otomatis tiap 60s. Manual refresh via reload page (F5).",
        },
        {
            "fitur": "Switch Active Kitchen",
            "path": "Sidebar > [dropdown 'Kitchen'] (di bawah brand)",
            "url": "(seluruh halaman)",
            "cara": (
                "1. Di sidebar, di bawah brand 'DAPUR PINTAR', ada label 'KITCHEN' uppercase kecil\n"
                "2. Dropdown <select> dengan list kitchen yang user punya akses\n"
                "3. Untuk superadmin: opsi pertama '🌐 All kitchens' → buka /admin/overview\n"
                "4. Pilih kitchen lain → halaman auto-reload, JWT rotated"
            ),
            "hasil": (
                "• POST /api/auth/switch-kitchen → terima token baru dengan active_kitchen_id beda\n"
                "• window.location.reload() → semua data scoped ulang\n"
                "• Header title berubah ke nama kitchen baru"
            ),
            "api": "POST /api/auth/switch-kitchen",
            "data": "Sandbox cuma 1 kitchen, dropdown nggak muncul (UI auto-hide)",
            "note": "Untuk test: bikin kitchen ke-2 via /admin/kitchens, dropdown akan muncul.",
        },
        {
            "fitur": "Toggle Dark Mode",
            "path": "Sidebar > [bottom panel] > ☀/☾ Dark mode",
            "url": "(seluruh halaman, persistent)",
            "cara": (
                "1. Scroll ke bawah sidebar\n"
                "2. Klik tombol '☾ Dark mode' (kalau lagi light) atau '☀ Light mode'\n"
                "3. Theme switch instan (CSS var swap)"
            ),
            "hasil": "Saved ke localStorage 'dark_mode'; apply ke seluruh page; persist next session.",
            "api": "(client-side only)",
            "data": "—",
            "note": "Tailwind CSS dark: variant; tidak hit backend.",
        },
        {
            "fitur": "Logout",
            "path": "Sidebar > [bottom panel] > ⏻ Logout",
            "url": "(seluruh halaman)",
            "cara": "1. Klik '⏻ Logout' di bottom sidebar.\n2. localStorage clear, redirect ke /login.",
            "hasil": "Token dihapus client-side. Backend nggak invalidate (stateless JWT) — token expire natural sesudah 7 hari.",
            "api": "(client-side only)",
            "data": "—",
            "note": "Buat full revoke pakai /api/auth/me + blacklist (tidak implement saat ini).",
        },
    ]),

    # ═════════════════════════════════════════════════════════════════════
    ("👑 PLATFORM_ADMIN — vendor SaaS / IT lintas yayasan", [
        {
            "fitur": "Manage Organizations",
            "path": "Sidebar > Platform > Organizations (🏢)",
            "url": "/admin/organizations",
            "cara": (
                "1. Sidebar section 'PLATFORM' — cuma muncul untuk role=platform_admin\n"
                "2. Klik 'Organizations'\n"
                "3. Tabel list org: Slug | Name | Active | Created\n"
                "4. Tombol 'Tambah Organisasi' → modal:\n"
                "   - Field 'Slug' (lowercase, unique, ex: 'yayasan-makmur')\n"
                "   - Field 'Name' (display name, ex: 'Yayasan Makmur Sehat')\n"
                "   - Toggle 'Active'\n"
                "5. Klik 'Save' → org muncul di tabel"
            ),
            "hasil": (
                "• Org row baru di tabel organizations\n"
                "• Bisa langsung di-assign kitchen (via /admin/kitchens dengan org_id terisi)\n"
                "• Login dengan username yang sama di >1 org butuh org_slug param disambiguation"
            ),
            "api": "GET/POST/PATCH/DELETE /api/organizations",
            "data": "Sandbox punya org: 'testsandbox' (id=19)",
            "note": "Hanya platform_admin yang bisa create/delete org. Superadmin tidak.",
        },
        {
            "fitur": "All Kitchens Cross-Org",
            "path": "Sidebar > Admin Dapur > All Kitchens (🌐)",
            "url": "/admin/overview",
            "cara": "1. Klik 'All Kitchens' (juga bisa dari kitchen dropdown → '🌐 All kitchens')\n2. Tabel: Org | Kitchen | Porsi Today | Compliance | Last Activity",
            "hasil": "View cross-org semua kitchen dalam satu tabel. Bisa drill-down per kitchen.",
            "api": "GET /api/admin/overview",
            "data": "Sandbox cuma 1 kitchen, jadi tabel cuma 1 baris",
            "note": "Superadmin lihat kitchen di org-nya saja; platform_admin lihat semua org.",
        },
    ]),

    # ═════════════════════════════════════════════════════════════════════
    ("🏛️ SUPERADMIN — yayasan owner (multi-SPPG dalam 1 org)", [
        {
            "fitur": "Manage Kitchens",
            "path": "Sidebar > Admin Dapur > Kitchens (🏠)",
            "url": "/admin/kitchens",
            "cara": (
                "1. Klik 'Kitchens' di sidebar section 'ADMIN DAPUR'\n"
                "2. Tabel: ID | Slug | Name | Printer | Active | Created\n"
                "3. Klik 'Tambah Kitchen' → modal form:\n"
                "   - 'Slug' (unique dalam org, ex: 'paseh-1')\n"
                "   - 'Name' (display, ex: 'DPMBG Paseh Pusat')\n"
                "   - 'Address'\n"
                "   - 'Printer Name' (nama device Windows)\n"
                "   - 'Printer Lang' (ZPL / TSPL)\n"
                "   - 'Label Title' (header label print)\n"
                "   - 'Timezone' (default Asia/Jakarta)\n"
                "4. Save → kitchen baru, scanner_key + cloud_print_key auto-generate (token random)\n"
                "5. Klik baris → modal detail dengan tombol 'Rotate Scanner Key' / 'Rotate Print Key'"
            ),
            "hasil": (
                "• Kitchen row di tabel kitchens, scoped ke org user\n"
                "• Scanner Android & cloud printer agent perlu key baru → setting di device-nya\n"
                "• Bisa di-assign user via /admin/users → tab 'Kitchens' (multi-select)"
            ),
            "api": "GET/POST/PATCH/DELETE /api/kitchens, POST /api/kitchens/{id}/rotate-scanner-key",
            "data": "Sandbox: 'Test Sandbox Kitchen' (id=78), printer 'TEST-PR1'",
            "note": "Delete = soft (active=false). Hard delete cuma platform_admin.",
        },
        {
            "fitur": "Multi-SPPG (Yayasan) View",
            "path": "Sidebar > Executive > [tab] Multi-SPPG (Yayasan)",
            "url": "/executive (tab=multi)",
            "cara": (
                "1. Buka 'Executive' → muncul tab di atas: 'Per Dapur' | 'Multi-SPPG (Yayasan)' | 'Platform (Cross-Org)'\n"
                "2. Klik tab 'Multi-SPPG (Yayasan)'\n"
                "3. Tabel ranking: Dapur | Porsi Target | Confirmed | Cost/Porsi | Defect % | Compliance\n"
                "4. Sort default: compliance score desc"
            ),
            "hasil": (
                "• Leaderboard kitchen dalam yayasan ini\n"
                "• GINI inequality coefficient di footer (0=equal, 1=skewed)\n"
                "• Kalau defect >5% di salah satu kitchen → highlight merah\n"
                "• Bisa drill-down klik baris → KPI per-kitchen"
            ),
            "api": "GET /api/executive/multi-kpi",
            "data": "Sandbox 1 kitchen → ranking trivially top-1",
            "note": "Tab cuma muncul kalau user superadmin atau platform_admin.",
        },
        {
            "fitur": "Compliance Bundle Export (Yayasan)",
            "path": "Sidebar > Executive > [tombol] 'Export Bundle'",
            "url": "/executive",
            "cara": (
                "1. Buka Executive Dashboard\n"
                "2. Tombol 'Export Compliance Bundle' (kanan atas)\n"
                "3. Modal: pilih 'Period Start' + 'Period End' (date range, biasanya 1 bulan)\n"
                "4. Pilih scope: 'This kitchen only' / 'All kitchens in org'\n"
                "5. Klik 'Generate' → progress bar (~5-10 detik)\n"
                "6. Browser download .zip"
            ),
            "hasil": (
                "• File 'compliance-bundle-{kitchen}-{period}.zip' download\n"
                "• Isi:\n"
                "  - lra_periods.csv (semua LRA biweekly signed)\n"
                "  - sample_retention.csv (sample 24-48 jam BGN)\n"
                "  - inspection_signoffs.csv (3-PIC signatures)\n"
                "  - aslap_reports.csv (weekly reports signed)\n"
                "  - menu_compliance.csv (AKG match per menu)\n"
                "  - certificates/ (PDF tanda tangan digital)\n"
                "• Audit-ready untuk BGN inspector"
            ),
            "api": "GET /api/executive/compliance-bundle?from=...&to=...",
            "data": "Sandbox punya 1 hari data, bundle bakal kecil — tambah re-seed beberapa hari",
            "note": "Endpoint streaming — file digenerate on-the-fly, tidak persist di server.",
        },
    ]),

    # ═════════════════════════════════════════════════════════════════════
    ("🧑‍💼 HEAD_SPPG / Kepala SPPG — admin 1 dapur (testadmin punya semua perm ini)", [
        {
            "fitur": "Approval Menu",
            "path": "Sidebar > Menu & Gizi > Approval Menu (✓)",
            "url": "/menu-approval",
            "cara": (
                "1. Klik 'Approval Menu' di sidebar section 'MENU & GIZI'\n"
                "2. Filter chip atas: 'all' / 'draft' / 'pending_review' / 'approved' / 'locked' / 'rejected' / 'archived'\n"
                "3. Default filter: pending_review (yang butuh approval)\n"
                "4. Tabel: Nama Menu | Tanggal Target | Source | Status | Cost/Porsi | Aksi\n"
                "5. Klik nama menu → detail panel kanan: items list, totals, AKG compare, biaya\n"
                "6. Tombol per status:\n"
                "   - Status draft: 'Submit for Review'\n"
                "   - Status pending_review: 'Approve' (hijau) / 'Reject' (merah, prompt alasan)\n"
                "   - Status approved: 'Lock' (gembok merah, freeze cycle)\n"
                "   - Status locked/rejected: 'Revert to Draft'\n"
                "   - Status approved/archived: 'Archive'"
            ),
            "hasil": (
                "• Status menu ber-update sesuai action\n"
                "• Notif auto-fire ke pembuat menu (Ahli Gizi)\n"
                "• approved_at + approved_by stamp di DB\n"
                "• Section bawah: Forecast Bahan dari Menu Approved (grouped by item)"
            ),
            "api": "POST /api/menu/saved/{id}/{submit|approve|reject|lock|revert-to-draft|archive}",
            "data": "Sandbox menu id=17 'Sandbox Menu 2026-05-10 — Nasi Ayam Bayam' status=approved",
            "note": "Approval cuma kasih notes optional, tidak ada workflow multi-step lain.",
        },
        {
            "fitur": "Forecast Bahan dari Menu Approved",
            "path": "Sidebar > Menu & Gizi > Approval Menu > [section bawah] Forecast Bahan",
            "url": "/menu-approval (scroll bottom)",
            "cara": (
                "1. Scroll ke bawah halaman Approval Menu\n"
                "2. Section 'Forecast Bahan dari Menu Approved'\n"
                "3. Date picker: 'From' + 'To' (ex: 1 minggu ke depan)\n"
                "4. Klik 'Run Forecast'\n"
                "5. Tabel hasil: Item | Total Grams | Container Estimate | Default Supplier"
            ),
            "hasil": "List bahan agregat untuk semua menu approved di range. Input langsung ke 'Generate PO from Forecast' (Akuntan).",
            "api": "GET /api/menu/forecast?from=...&to=...",
            "data": "Sandbox: 1 menu approved → forecast = 600g beras + 400g ayam + 250g bayam (untuk 5 porsi)",
            "note": "Hanya menu status approved/locked yang dihitung; draft & rejected diabaikan.",
        },
        {
            "fitur": "Joint Inspection — Sign-off Quality",
            "path": "Sidebar > Penerimaan Bahan > Joint Inspection (🤝) > [klik inspection] > [section] '3 Sign-Off' > role 'quality'",
            "url": "/inspections",
            "cara": (
                "1. Klik 'Joint Inspection' di sidebar\n"
                "2. Tabel inspection list, default filter 'pending'\n"
                "3. Klik 'Buka' di baris inspection\n"
                "4. Modal/panel detail: header 'INS-{id} · PO-{po_id}', status, lines table\n"
                "5. Section '3 Sign-Off' di tengah, ada 3 baris role: quality / quantity / physical\n"
                "6. Di baris 'quality': 2 tombol\n"
                "   - 'Approve' (hijau) → modal canvas tanda tangan + notes optional\n"
                "   - 'Reject' (merah) → modal alasan reject + severity\n"
                "7. Tanda tangan digital → save (canvas → base64 PNG)"
            ),
            "hasil": (
                "• Row inspection_signoffs INSERT: role=quality, user_id=lu, signed_at=now, status=approved\n"
                "• Tombol Approve/Reject quality jadi disabled (sudah signed)\n"
                "• Setelah ke-3 role signed → tombol 'Finalize' muncul di footer modal"
            ),
            "api": "POST /api/inspections/{id}/signoff (body: role=quality, status=approved)",
            "data": "Sandbox: inspection id=13 sudah 3 sign-offs lengkap (quality+quantity+physical)",
            "note": "Quality TTD bisa juga dilakukan Ahli Gizi. Head SPPG override semua role.",
        },
        {
            "fitur": "Joint Inspection — Reject Bahan + Auto-Dispute",
            "path": "Joint Inspection > [klik inspection] > line item > tombol 'Reject'",
            "url": "/inspections",
            "cara": (
                "1. Buka inspection detail\n"
                "2. Tabel lines: setiap baris ada 'Accept' / 'Reject' button (kalau status pending)\n"
                "3. Klik 'Reject' di line item rusak\n"
                "4. Prompt: 'Alasan reject?' + dropdown severity (low / medium / high)\n"
                "5. Submit"
            ),
            "hasil": (
                "• Line status → 'rejected'\n"
                "• Row supplier_disputes INSERT (severity, reason, photo_path null kalau ga upload)\n"
                "• Supplier rating decrement -1 (5 → 4 → 3...)\n"
                "• Notif ke Akuntan untuk follow-up dispute resolution\n"
                "• PO status berubah 'partial' kalau line lain accepted, 'rejected' kalau semua lines rejected"
            ),
            "api": "POST /api/inspections/{id}/lines/{line_id}/reject",
            "data": "Sandbox: belum ada dispute (semua line accepted)",
            "note": "Reject di Inspection berbeda dari Reject di Menu Approval. Ini scope bahan fisik, bukan menu plan.",
        },
        {
            "fitur": "Joint Inspection — Finalize",
            "path": "Joint Inspection > [klik inspection] > footer 'Finalize'",
            "url": "/inspections",
            "cara": "1. Buka inspection setelah 3 sign-off lengkap.\n2. Tombol 'Finalize' aktif di footer modal (kanan).\n3. Klik → confirm.",
            "hasil": (
                "• Status inspection → 'accepted' (semua lines accepted) / 'partial' (sebagian) / 'rejected' (semua reject)\n"
                "• completed_at = now()\n"
                "• Bahan accepted masuk ke tabel items dengan id 'BHN-XXXXX' (FIFO routing)\n"
                "• Notif ke Kepala Chef: 'Bahan ready, batch bisa dimulai'"
            ),
            "api": "POST /api/inspections/{id}/finalize",
            "data": "Sandbox: inspection id=13 sudah finalized (status=accepted)",
            "note": "Setelah finalize, sign-off ga bisa di-edit lagi.",
        },
        {
            "fitur": "LRA Sign-off (Biweekly)",
            "path": "Sidebar > Keuangan > Akuntan Finance (💰) > [tab] LRA Biweekly",
            "url": "/finance (tab=lra)",
            "cara": (
                "1. Buka 'Akuntan Finance' di sidebar\n"
                "2. Klik tab 'LRA Biweekly' (5th tab dari kiri)\n"
                "3. Tabel period list: Period | Start | End | Total Expense | Status (draft/submitted/signed)\n"
                "4. Klik period status='submitted' (dari Akuntan)\n"
                "5. Modal review: breakdown expense per kategori, total budget vs realisasi\n"
                "6. Klik 'Sign-off' → confirm\n"
            ),
            "hasil": (
                "• Status period → 'signed'\n"
                "• signed_at + signed_by stamp\n"
                "• Period auto-include di Compliance Bundle berikutnya\n"
                "• Notif ke Akuntan: 'LRA signed, ready for BGN submission'"
            ),
            "api": "POST /api/finance/lra/periods/{period_id}/submit",
            "data": "Sandbox: belum ada LRA (perlu generate dulu — lihat fitur Akuntan)",
            "note": "Sign-off di-trigger Kepala SPPG only. Akuntan generate, tapi tidak bisa sign sendiri.",
        },
        {
            "fitur": "Manajemen Pengguna",
            "path": "Sidebar > Admin Dapur > Users (👤)",
            "url": "/admin/users",
            "cara": (
                "1. Klik 'Users' di sidebar section 'ADMIN DAPUR'\n"
                "2. Tabel: ID | Username | Role | Org | Kitchens (jumlah) | Created\n"
                "3. Tombol 'Tambah User' → modal:\n"
                "   - 'Username' (lowercase, ex: 'nutri1')\n"
                "   - 'Password' (min 8 char)\n"
                "   - 'Role' dropdown: user / superadmin / platform_admin\n"
                "   - 'Org' (auto-fill org user yang login, kecuali platform_admin bisa pilih)\n"
                "4. Save → user muncul di tabel\n"
                "5. Klik baris user → modal detail dengan tab 'Kitchens':\n"
                "   - 'Tambah Kitchen' → pilih kitchen + role kitchen-level (head_sppg/nutritionist/accountant/aslap/head_kitchen)\n"
                "6. Tombol 'Reset Password' di header modal → input password baru → Save\n"
                "7. Tombol 'Hapus' → soft delete (user nonaktif)"
            ),
            "hasil": (
                "• User CRUD lengkap; password bcrypt-hashed\n"
                "• User-Kitchen mapping di tabel user_kitchens (many-to-many)\n"
                "• User langsung bisa login dengan kredensial baru\n"
                "• Audit log INSERT untuk setiap operasi (user.create, user.update, user.password_reset, user.delete)"
            ),
            "api": "GET/POST/PATCH/DELETE /api/users, POST /api/users/{id}/kitchens",
            "data": "Sandbox: 1 user (testadmin id=246, role=superadmin)",
            "note": "Buat test role berbeda — bikin nutri1, akun1, aslap1, chef1 → assign kitchen role di tab Kitchens.",
        },
        {
            "fitur": "Manajemen Sekolah",
            "path": "Sidebar > Master Data > Sekolah (🏫)",
            "url": "/admin/schools",
            "cara": (
                "1. Klik 'Sekolah' di sidebar section 'MASTER DATA'\n"
                "2. Tabel: ID | Nama | Level | Age Group | Siswa | Distance (m) | Active | Aksi\n"
                "3. Tombol 'Tambah Sekolah' → form:\n"
                "   - 'Nama' (ex: 'SDN 1 Sukamanah')\n"
                "   - 'Address'\n"
                "   - 'Level' dropdown: PAUD / TK / SD / SMP / SMA\n"
                "   - 'Age Group' (ex: 'SD (7-9 tahun)' — drives AKG preset)\n"
                "   - 'Student Count' (jumlah siswa)\n"
                "   - 'Distance' (meter dari kitchen)\n"
                "   - 'GPS Lat' + 'GPS Long' (optional)\n"
                "   - 'Contact' (HP guru)\n"
                "4. Save → sekolah ditambahkan, is_active=true\n"
                "5. Edit: klik baris → form pre-fill, ubah, save\n"
                "6. Hapus: klik 'Delete' di kolom Aksi → soft delete (is_active=false)"
            ),
            "hasil": (
                "• Sekolah jadi target distribusi\n"
                "• Wave classifier auto: PAUD/TK/SD-low → wave 1 pagi; SD-high/SMP/SMA → wave 2 siang\n"
                "• Distance dipakai routing engine sopir\n"
                "• Age Group → AKG preset di Build Menu Manual"
            ),
            "api": "GET /api/admin/schools, POST/PATCH/DELETE /api/admin/schools/{id}",
            "data": "Sandbox: 3 sekolah (TK Mawar, SD Tunas Bangsa, SMP Bina)",
            "note": "Wave classifier hard-coded di backend by age_group prefix. Override perlu code change.",
        },
        {
            "fitur": "Manajemen Supplier",
            "path": "Sidebar > Master Data > Supplier (📦)",
            "url": "/admin/suppliers",
            "cara": (
                "1. Klik 'Supplier' di sidebar section 'MASTER DATA'\n"
                "2. Tabel: ID | Nama | Kategori | Rating ⭐ | Active | Aksi\n"
                "3. Tombol 'Tambah Supplier' → form:\n"
                "   - 'Nama' (ex: 'Toko Sayur Bu Tini')\n"
                "   - 'Contact' (HP / WA)\n"
                "   - 'NPWP' (optional)\n"
                "   - 'Rekening' (no rek bank)\n"
                "   - 'Bank Name' (ex: 'BCA')\n"
                "   - 'Kategori' (ex: 'sayur', 'daging', 'beras', 'bumbu')\n"
                "   - 'Rating' (1-5, default 5)\n"
                "   - 'Notes'\n"
                "4. Save → supplier siap untuk PO\n"
                "5. Rating auto-decrement -1 setiap dispute (di Inspection reject)\n"
                "6. Klik baris → detail + dispute history"
            ),
            "hasil": "Supplier list aktif; rating dipakai untuk recommend di PO Generator.",
            "api": "GET/POST/PATCH/DELETE /api/suppliers",
            "data": "Sandbox: 3 supplier (Toko Sayur ⭐5, Pak Budi Daging ⭐5, Beras Sehat ⭐4)",
            "note": "Rating <2 → supplier highlight merah, manual review.",
        },
        {
            "fitur": "Resolve Permintaan Khusus Siswa",
            "path": "Sidebar > Menu & Gizi > Approval Menu > [tab/section] Permintaan Khusus",
            "url": "/menu-approval (section)",
            "cara": (
                "1. Buka 'Approval Menu'\n"
                "2. Section 'Permintaan Khusus Siswa' di kanan atau scroll bawah\n"
                "3. Filter status: 'open' / 'confirmed' / 'rejected' / 'fulfilled'\n"
                "4. Klik request → detail: nama siswa, kelas, sekolah, request_text\n"
                "5. Tombol:\n"
                "   - 'Confirm' (terima, akan diakomodasi di menu)\n"
                "   - 'Reject' (tidak bisa, tulis alasan)\n"
                "   - 'Fulfilled' (sudah dilayani)"
            ),
            "hasil": "Status request update; ahli_gizi_notes terisi; notif ke pembuat (ASLAP / Ahli Gizi).",
            "api": "PATCH /api/student-requests/{id}",
            "data": "Sandbox: 1 request 'Anak alergi telur, butuh menu pengganti' (status=open)",
            "note": "Capture request bisa Ahli Gizi atau ASLAP; resolve cuma head_sppg / Ahli Gizi.",
        },
        {
            "fitur": "ASLAP Weekly Report Sign-off",
            "path": "Sidebar > Lapangan & Monitoring > ASLAP Daily (📝) > [tab] Weekly Report",
            "url": "/aslap (tab=reports)",
            "cara": (
                "1. Klik 'ASLAP Daily' di sidebar\n"
                "2. Tab paling kanan: 'Weekly Report'\n"
                "3. Tabel report: Week | Status (pending_signoff / signed) | Submitted By\n"
                "4. Klik report status='pending_signoff' → preview content (checklist summary, water quality, observations, comm logs)\n"
                "5. Tombol 'Sign-off' → confirm"
            ),
            "hasil": "Status report → 'signed'. Masuk Compliance Bundle. Audit-ready.",
            "api": "POST /api/aslap/reports/{id}/submit",
            "data": "Sandbox: ASLAP report belum di-generate (lihat fitur ASLAP)",
            "note": "Generate dilakukan ASLAP, sign-off cuma Kepala SPPG.",
        },
        {
            "fitur": "Override Harga Bahan",
            "path": "Sidebar > Keuangan > Akuntan Finance > [tab] Price Trends > [klik bahan] > 'Override'",
            "url": "/finance (tab=trends)",
            "cara": (
                "1. Buka 'Akuntan Finance' → tab 'Price Trends'\n"
                "2. Tabel bahan: Code | Nama | Harga Scrape | Manual Override | Source\n"
                "3. Klik bahan → modal detail (chart 30 hari + history)\n"
                "4. Tombol 'Override Manual Price' → form:\n"
                "   - 'Manual Price' (Rp per 100g)\n"
                "   - 'Source' (ex: 'Invoice Bu Tini 2026-05')\n"
                "5. Save"
            ),
            "hasil": (
                "• food_prices.manual_price terisi → menang dari price_per_100g (scrape)\n"
                "• Row baru di food_prices_history (action=manual)\n"
                "• Menu calc pakai harga manual sampai di-clear\n"
                "• Audit log INSERT (prices.override)"
            ),
            "api": "PATCH /api/menu/prices/{food_code}",
            "data": "Sandbox: belum ada manual override (semua pakai seed price)",
            "note": "Clear override: PATCH dengan body {manual_price: null}.",
        },
    ]),

    # ═════════════════════════════════════════════════════════════════════
    ("🥗 NUTRITIONIST / Ahli Gizi", [
        {
            "fitur": "Build Menu Manual",
            "path": "Sidebar > Menu & Gizi > Build Manual (✎)",
            "url": "/menu-manual",
            "cara": (
                "1. Klik 'Build Manual' di sidebar\n"
                "2. Halaman split 2 kolom:\n"
                "   - Kiri: Katalog Bahan (search + filter category)\n"
                "   - Kanan: Menu Builder (drop zone)\n"
                "3. Pilih 'Age Group' di dropdown atas: TK 4-6 / SD 7-9 / SD 10-12 / SMP / SMA\n"
                "4. Drag bahan dari kiri ke kanan, atau klik '+' di card bahan\n"
                "5. Setiap bahan di kanan ada input gram → ketik jumlah (ex: 120g beras)\n"
                "6. Real-time update: 'Totals' card (energy, protein, fat, carbs)\n"
                "7. 'AKG Compare' card warna-coded:\n"
                "   - Hijau ✅ kalau total ≥ AKG min\n"
                "   - Kuning ⚠️ kalau 80-100% AKG\n"
                "   - Merah ⚠️ kalau <80%\n"
                "8. 'Cost per Porsi' card (auto Rp)\n"
                "9. Tombol 'Save Draft' → modal:\n"
                "   - 'Nama Menu' (ex: 'SD Senin Nasi Ayam Bayam')\n"
                "   - 'Target Date' (kapan menu ini disajikan)\n"
                "   - 'Target School' (optional, kalau scoped ke sekolah tertentu)\n"
                "10. Save → status='draft'"
            ),
            "hasil": (
                "• Row saved_menus INSERT, payload JSON {items, totals, cost}\n"
                "• Status 'draft' — perlu submit untuk approval\n"
                "• Bahan harus punya price (cek di Master Data > Food Prices)\n"
                "• AKG compare pakai TKPI nutrient table + Permenkes preset"
            ),
            "api": "POST /api/menu/calc (real-time), POST /api/menu/saved (final)",
            "data": "Sandbox menu (id=17): 120g beras + 80g ayam + 50g bayam, target_porsi=5, status=approved",
            "note": "Drag-drop pakai react-dnd. Mobile-friendly via touch events.",
        },
        {
            "fitur": "Reverse Optimizer (Auto-Compose)",
            "path": "Sidebar > Menu & Gizi > Build Manual > [tombol] '⚡ Auto-Optimize'",
            "url": "/menu-manual",
            "cara": (
                "1. Di halaman 'Build Manual', tombol '⚡ Auto-Optimize' di header\n"
                "2. Modal:\n"
                "   - 'Target Energy' (kcal min, default ngikut age_group)\n"
                "   - 'Target Protein' (g)\n"
                "   - 'Budget per Porsi' (Rp max)\n"
                "   - 'Required Items' (multi-select, harus muncul, ex: beras+protein)\n"
                "   - 'Forbidden Items' (multi-select, jangan dipilih, ex: bahan alergi)\n"
                "3. Klik 'Generate' → loading 5-10 detik (linear programming solver)\n"
                "4. Hasil: list kombinasi terbaik (sorted by cost asc, AKG match desc)\n"
                "5. Klik 'Apply' → menu builder kanan auto-fill"
            ),
            "hasil": "Menu draft otomatis ngikutin constraint. Bisa di-edit manual lagi sebelum save.",
            "api": "POST /api/menu/optimize",
            "data": "Sandbox belum pernah pakai optimizer (manual menu only)",
            "note": "Solver pakai PuLP di backend. Slow kalau katalog >100 bahan.",
        },
        {
            "fitur": "Cek Siklus 20 Hari (Anti-Bosen)",
            "path": "Sidebar > Menu & Gizi > Build Manual > [tombol] '🔄 Cek Siklus'",
            "url": "/menu-manual",
            "cara": (
                "1. Tombol '🔄 Cek Siklus' di header halaman\n"
                "2. Modal: tampilkan menu approved/locked dalam 20 hari terakhir\n"
                "3. Tabel: Item | Frekuensi (kali muncul) | Limit BGN | Status\n"
                "4. Warning kalau ada item lewat limit (ex: ayam max 6x/20hari)"
            ),
            "hasil": "Insight repetisi bahan; help Ahli Gizi vary menu sesuai pedoman BGN.",
            "api": "GET /api/menu/cycle-check?days=20",
            "data": "Sandbox: cuma 1 menu approved → cycle ratio rendah, semua hijau",
            "note": "Limit BGN configurable di backend constants. Default conservative.",
        },
        {
            "fitur": "Forecast Kebutuhan Bahan",
            "path": "Sidebar > Menu & Gizi > Build Manual > [tombol] 'Forecast' / atau di Approval Menu bottom",
            "url": "/menu-manual atau /menu-approval",
            "cara": "1. Tombol 'Forecast' di header.\n2. Date range picker.\n3. Klik 'Run' → tabel: Item | Total Grams | Container Estimate | Default Supplier.",
            "hasil": "Daftar agregat bahan untuk semua menu approved/locked di range. Input untuk PO Akuntan.",
            "api": "GET /api/menu/forecast?from=...&to=...",
            "data": "Sandbox: 1 menu × 5 porsi = 600g beras + 400g ayam + 250g bayam",
            "note": "Hanya menu approved/locked yang dihitung. Status draft/rejected diabaikan.",
        },
        {
            "fitur": "Submit Menu for Review",
            "path": "Build Manual > [setelah save draft] tombol 'Submit for Review' / atau di Approval Menu",
            "url": "/menu-manual atau /menu-approval",
            "cara": "1. Pilih menu draft.\n2. Klik 'Submit for Review' → confirm.",
            "hasil": "Status → 'pending_review'. Notif menu.pending_review ke head_sppg.",
            "api": "POST /api/menu/saved/{id}/submit",
            "data": "Sandbox menu sudah lewat fase ini",
            "note": "Hanya pembuat (created_by) yang bisa submit menu draft.",
        },
        {
            "fitur": "Laporan Gizi Harian (Nutrisi Harian)",
            "path": "Sidebar > Menu & Gizi > Nutrisi Harian (📊)",
            "url": "/nutrisi",
            "cara": (
                "1. Klik 'Nutrisi Harian' di sidebar\n"
                "2. Date picker: pilih tanggal\n"
                "3. Tabel/chart: kompliance AKG per menu hari itu\n"
                "4. Section: protein/energy/fat/carbs per menu vs target\n"
                "5. Export PDF/CSV (kalau tombol tersedia)"
            ),
            "hasil": "Snapshot harian compliance gizi. Pakai untuk laporan ke supervisor BGN.",
            "api": "GET /api/nutrition/report?date=...",
            "data": "Sandbox: menu approved hari ini → ada data",
            "note": "Hanya menu yang sudah approved/locked masuk laporan.",
        },
        {
            "fitur": "QC Approve Produksi (Quality Control)",
            "path": "Sidebar > Produksi & Distribusi > Production (🍳) > [batch status qc_pending] tombol 'QC'",
            "url": "/production",
            "cara": (
                "1. Klik 'Production' di sidebar\n"
                "2. Section 'Batch Aktif' — list batch dengan status (started/qc_pending/qc_passed/ended/aborted)\n"
                "3. Klik batch status='qc_pending'\n"
                "4. Modal detail: menu_name, target_porsi, started_at, elapsed_minutes\n"
                "5. Tombol hijau 'QC' → modal:\n"
                "   - 'Sample Location' (ex: 'Kulkas QC rak 2')\n"
                "   - 'Notes' (ex: 'Warna OK, aroma sedap')\n"
                "6. Save"
            ),
            "hasil": (
                "• Batch status → 'qc_passed'\n"
                "• Sample row INSERT di production_samples (expire_at = now + 48 jam)\n"
                "• Notif ke ASLAP: 'Batch ready, siap dispatch distribusi'"
            ),
            "api": "POST /api/production/batches/{id}/qc",
            "data": "Sandbox batch id=2 sudah qc_passed",
            "note": "Sample wajib BGN — disimpan 24-48 jam untuk inspeksi.",
        },
        {
            "fitur": "Manajemen Sampel Makanan",
            "path": "Sidebar > Produksi & Distribusi > Production > [bagian batch detail] Sampel section",
            "url": "/production",
            "cara": "1. Buka batch detail (status qc_passed/ended).\n2. Section 'Sampel' menampilkan list sample + countdown expire.\n3. Tombol 'Mark Discarded' kalau sudah lewat 48 jam.",
            "hasil": "Sample tracking; expired samples flagged untuk discard.",
            "api": "GET /api/production/batches/{id}/samples, POST /samples/{id}/discard",
            "data": "Sandbox: 1 sample dari batch id=2",
            "note": "Sample location custom string. Foto sample optional (upload via separate endpoint).",
        },
    ]),

    # ═════════════════════════════════════════════════════════════════════
    ("💰 ACCOUNTANT / Akuntan", [
        {
            "fitur": "Purchase Orders — Manual",
            "path": "Sidebar > Penerimaan Bahan > Purchase Orders (📋) > [tombol] 'PO Baru'",
            "url": "/purchase-orders",
            "cara": (
                "1. Klik 'Purchase Orders' di sidebar\n"
                "2. Tabel PO: ID | Tanggal | Supplier | Lines | Total IDR | Status | Aksi\n"
                "3. Filter chip: all / draft / sent / partial / received / closed\n"
                "4. Tombol 'PO Baru' → form muncul:\n"
                "   - 'Supplier' dropdown (dari Master Data)\n"
                "   - 'Expected Delivery Date'\n"
                "   - 'Notes'\n"
                "   - Section 'Lines':\n"
                "     · Tombol '+ Tambah Line'\n"
                "     · Per line: Item Name | Item Code | Total Weight (g) | Unit | Containers | Unit Price (Rp/kg) | Line Total\n"
                "     · Line total auto-calc: weight/1000 × unit_price\n"
                "     · Tombol '×' merah untuk hapus line\n"
                "5. Total auto-sum di footer\n"
                "6. Klik 'Save' → PO status='draft'\n"
                "7. Di tabel, baris PO draft ada tombol 'Kirim' → status='sent' (notif supplier)\n"
                "8. Tombol 'Hapus' (cuma untuk status=draft)\n"
                "9. Tombol 'Detail' buka modal full info"
            ),
            "hasil": (
                "• PO row + lines INSERT\n"
                "• total_amount_idr auto-sum dari lines\n"
                "• Status flow: draft → sent → partial/received → closed\n"
                "• Notif ke ASLAP saat status='sent' (siap inspect saat truk datang)"
            ),
            "api": "GET/POST/PATCH/DELETE /api/purchase-orders, POST /lines",
            "data": "Sandbox PO id=10 status=sent (ayam 8kg + bayam 5kg + beras 12kg, total ~Rp 600rb)",
            "note": "PO draft bisa di-edit; sent ke atas read-only kecuali admin override.",
        },
        {
            "fitur": "Auto-Generate PO from Forecast",
            "path": "Sidebar > Keuangan > Akuntan Finance > [tab] PO Generator",
            "url": "/finance (tab=po)",
            "cara": (
                "1. Buka 'Akuntan Finance' → tab 'PO Generator' (paling kanan)\n"
                "2. Form:\n"
                "   - 'From Date' / 'To Date' (range menu approved)\n"
                "   - 'Supplier' (default per kategori, atau pilih satu untuk semua)\n"
                "   - 'Notes'\n"
                "3. Klik 'Generate' → preview PO (lines auto-fill dari forecast)\n"
                "4. Edit kalau perlu (ubah qty / supplier per line)\n"
                "5. Save → PO created status='draft'\n"
                "6. Lanjut ke /purchase-orders → klik 'Kirim'"
            ),
            "hasil": "PO auto dengan lines = forecast Ahli Gizi. Skip manual entry.",
            "api": "POST /api/finance/po/generate-from-forecast",
            "data": "Sandbox: PO id=10 dibuat manual seed (bukan dari forecast)",
            "note": "Forecast ngikut menu approved/locked. Default supplier dari item.default_supplier_id.",
        },
        {
            "fitur": "Cost-per-porsi",
            "path": "Sidebar > Keuangan > Akuntan Finance > [tab] Cost-per-porsi",
            "url": "/finance (tab=cost)",
            "cara": "1. Tab 'Cost-per-porsi' (default tab Finance).\n2. Date picker.\n3. Tabel: Tanggal | Menu | Bahan Cost | Overhead | Total Cost | Porsi | Cost/Porsi.",
            "hasil": "Cost breakdown per porsi per hari. Pakai untuk LRA realisasi vs RAB.",
            "api": "GET /api/finance/cost-per-porsi",
            "data": "Sandbox: 1 batch hari ini, cost ~Rp 1.100/porsi",
            "note": "Overhead = expense kategori utility/transport/maintenance dibagi rata porsi hari itu.",
        },
        {
            "fitur": "Price Trends (Tren Harga)",
            "path": "Sidebar > Keuangan > Akuntan Finance > [tab] Price Trends",
            "url": "/finance (tab=trends)",
            "cara": (
                "1. Tab 'Price Trends'\n"
                "2. Tabel bahan: Code | Nama | Harga Sekarang | 7d Avg | 30d Avg | Trend (↑/↓/→)\n"
                "3. Klik bahan → modal chart 30 hari (line chart) + table history\n"
                "4. Bisa override manual price dari sini (lihat fitur head_sppg Override Harga)"
            ),
            "hasil": "Insight pergerakan harga; bantu negotiate supplier kalau bahan naik konsisten.",
            "api": "GET /api/finance/price-trends, /summary",
            "data": "Sandbox: 8 bahan seed, harga static (belum ada history)",
            "note": "Daily price scraper jalan via APScheduler — populate price history otomatis.",
        },
        {
            "fitur": "Spike Alerts",
            "path": "Sidebar > Keuangan > Akuntan Finance > [tab] Spike Alerts",
            "url": "/finance (tab=alerts)",
            "cara": "1. Tab 'Spike Alerts'.\n2. Tabel bahan dengan harga naik >20% dari 7-day moving avg.\n3. Severity chip: low (20-30%) / med (30-50%) / high (>50%).",
            "hasil": "Daftar bahan urgent untuk re-negotiate / cari supplier alternatif.",
            "api": "GET /api/finance/spike-alerts",
            "data": "Sandbox: 0 alert (harga belum berubah dari seed)",
            "note": "Threshold spike dapat di-configure via env var.",
        },
        {
            "fitur": "Expense Tracker",
            "path": "Sidebar > Keuangan > Akuntan Finance > [tab] Expense",
            "url": "/finance (tab=expense)",
            "cara": (
                "1. Tab 'Expense'\n"
                "2. Tabel: Tanggal | Kategori | Amount | Description | Created By\n"
                "3. Tombol 'Tambah Expense' → form:\n"
                "   - 'Date' (default today)\n"
                "   - 'Category' dropdown: utility / transport / maintenance / other\n"
                "   - 'Amount' (Rp)\n"
                "   - 'Description' (free text)\n"
                "4. Save\n"
                "5. Hapus: tombol '×' di kolom aksi"
            ),
            "hasil": "Expense log harian. Aggregate masuk LRA biweekly.",
            "api": "POST /api/finance/expenses, DELETE /{id}",
            "data": "Sandbox: 2 expense per hari (PLN Rp 250rb + transport Rp 120rb)",
            "note": "Hard delete tersedia (Akuntan & Head SPPG). Audit log INSERT setiap delete.",
        },
        {
            "fitur": "LRA Generate Biweekly",
            "path": "Sidebar > Keuangan > Akuntan Finance > [tab] LRA Biweekly > [tombol] 'Generate Period'",
            "url": "/finance (tab=lra)",
            "cara": (
                "1. Tab 'LRA Biweekly'\n"
                "2. Tabel period list (kalau ada period sebelumnya)\n"
                "3. Tombol 'Generate New Period' → form:\n"
                "   - 'Start Date' + 'End Date' (biasanya 14 hari)\n"
                "   - 'Notes'\n"
                "4. Klik 'Generate' → backend aggregate semua expense + cost-per-porsi di range\n"
                "5. Hasil: period status='draft' → klik 'Submit' → status='submitted' (siap di-sign Kepala SPPG)"
            ),
            "hasil": (
                "• lra_periods row INSERT\n"
                "• Breakdown: total expense, total porsi, realisasi vs target\n"
                "• PDF preview (kalau ada generator)\n"
                "• Notif ke head_sppg untuk sign-off"
            ),
            "api": "POST /api/finance/lra/generate, POST /periods/{id}/submit",
            "data": "Sandbox: belum ada period (perlu manual generate)",
            "note": "Hanya Akuntan / head_sppg yang bisa generate. Period overlap di-prevent (unique constraint).",
        },
        {
            "fitur": "Pembayaran Relawan",
            "path": "Sidebar > Keuangan > Akuntan Finance > [bagian] Volunteer Payments",
            "url": "/finance (di tab expense atau dedicated section)",
            "cara": (
                "1. Section 'Volunteer Payments'\n"
                "2. Tombol 'Tambah' → form:\n"
                "   - 'Name'\n"
                "   - 'Role' (ibu_dapur / driver / lain-lain)\n"
                "   - 'Honor' (Rp)\n"
                "   - 'Date'\n"
                "3. Save"
            ),
            "hasil": "Payment log; aggregate masuk LRA (kategori 'volunteer').",
            "api": "POST /api/finance/volunteers",
            "data": "Sandbox: 1 volunteer 'Bu Yani' Rp 75rb per hari",
            "note": "Tidak ada payroll pipeline; cuma log expense.",
        },
        {
            "fitur": "Joint Inspection — Sign-off Quantity",
            "path": "Sidebar > Penerimaan Bahan > Joint Inspection > [klik inspection] > role 'quantity' tombol 'Approve'",
            "url": "/inspections",
            "cara": "1. Buka inspection.\n2. Section '3 Sign-Off' baris 'quantity'.\n3. Klik 'Approve' (hijau) → tanda tangan canvas → submit.",
            "hasil": "TTD quantity recorded. Akuntan verify weight & count match PO.",
            "api": "POST /api/inspections/{id}/signoff (role=quantity)",
            "data": "Sandbox inspection id=13 sudah quantity-signed",
            "note": "Quantity TTD spesifik ke role akuntan/head_sppg. Tidak bisa di-do oleh aslap/nutritionist.",
        },
        {
            "fitur": "Manage Vehicles & Drivers",
            "path": "Sidebar > Produksi & Distribusi > Distribusi (🚐) > [tab] Vehicle & Driver",
            "url": "/distributions (tab=fleet)",
            "cara": (
                "1. Buka 'Distribusi'\n"
                "2. Tab 'Vehicle & Driver'\n"
                "3. Section '🚐 Kendaraan': tombol '+ Tambah' → name + plate + capacity\n"
                "4. Section '👤 Driver': tombol '+ Tambah' → name + phone + license_no\n"
                "5. Tombol '×' untuk delete masing-masing"
            ),
            "hasil": "Master vehicle/driver. Dipakai dispatch distribusi.",
            "api": "GET/POST/DELETE /api/vehicles, /drivers",
            "data": "Sandbox: belum ada vehicle/driver (optional)",
            "note": "Wajib head_sppg / accountant. Aslap cuma view.",
        },
        {
            "fitur": "Resolve Dispute Supplier",
            "path": "Sidebar > Penerimaan Bahan > Joint Inspection > [tab/list] Disputes",
            "url": "/inspections (atau dedicated /disputes)",
            "cara": (
                "1. Section disputes (kalau ada baris dengan status 'open')\n"
                "2. Klik dispute → modal:\n"
                "   - Detail: item_name, reason, severity, supplier, photo\n"
                "3. Tombol 'Resolve' → form:\n"
                "   - 'Resolution Notes' (ex: 'Supplier ganti produk; refund parsial')\n"
                "4. Save"
            ),
            "hasil": "Status → 'resolved'. Audit trail lengkap.",
            "api": "POST /api/disputes/{id}/resolve",
            "data": "Sandbox: 0 dispute (tidak ada line yang di-reject)",
            "note": "Severity 'high' auto-notify head_sppg + pause supplier (status active=false).",
        },
    ]),

    # ═════════════════════════════════════════════════════════════════════
    ("🦺 ASLAP / Asisten Lapangan", [
        {
            "fitur": "Buat Inspection Baru",
            "path": "Sidebar > Penerimaan Bahan > Joint Inspection > [tombol] 'Inspeksi Baru'",
            "url": "/inspections",
            "cara": (
                "1. Klik 'Joint Inspection'\n"
                "2. Tombol 'Inspeksi Baru' di header\n"
                "3. Form muncul:\n"
                "   - 'PO ID' dropdown (PO status='sent' yang belum ada inspection)\n"
                "   - 'Notes' (ex: 'Truk dateng 04:00')\n"
                "4. Klik 'Buat'\n"
                "5. Inspection muncul status='pending', lines auto-load dari PO"
            ),
            "hasil": "Inspection ready untuk 3 sign-off. Notif ke quality/quantity/physical PIC.",
            "api": "POST /api/inspections (body: po_id, notes)",
            "data": "Sandbox inspection id=13 sudah finalized",
            "note": "Bisa juga inspection tanpa PO (manual receipt) — po_id=null.",
        },
        {
            "fitur": "Container Split + Multi-Label Print",
            "path": "Joint Inspection > [klik inspection] > line item > tombol 'Accept'",
            "url": "/inspections",
            "cara": (
                "1. Buka inspection setelah 3 sign-off\n"
                "2. Tabel lines, klik tombol 'Accept' di line\n"
                "3. Modal:\n"
                "   - 'Containers': list of {weight_grams}\n"
                "     · Default: 1 container = full expected weight\n"
                "     · Tombol '+ Split' → tambah container\n"
                "     · Set weight per container (auto-distribute atau manual)\n"
                "     · Ex: 100kg ayam → 10 box × 10kg + 1 box leftover\n"
                "   - 'Storage Routing' dropdown: cook_immediate / refrigerate / freeze\n"
                "   - 'Notes'\n"
                "4. Klik 'Accept' → backend generate N container ID 'BHN-XXXXXXXX'\n"
                "5. Print queue ter-add N label (sesuai container count)"
            ),
            "hasil": (
                "• Items table INSERT N rows (1 per container)\n"
                "• Print jobs queued (cetak label thermal)\n"
                "• Storage routing menentukan suhu penyimpanan\n"
                "• Line status → 'accepted'"
            ),
            "api": "POST /api/inspections/{id}/lines/{line_id}/accept",
            "data": "Sandbox: 3 lines accepted, 1 container per line (no split)",
            "note": "Container ID 'BHN-XXXXXXXX' = 8-char alphanumeric. Unique per kitchen.",
        },
        {
            "fitur": "Joint Inspection — Sign-off Physical",
            "path": "Joint Inspection > [klik inspection] > role 'physical' tombol 'Approve'",
            "url": "/inspections",
            "cara": "1. Buka inspection.\n2. Section sign-off, baris 'physical'.\n3. Klik 'Approve' → canvas tanda tangan → submit.",
            "hasil": "TTD physical (kondisi fisik kemasan + cold-chain). Permission check: aslap atau head_sppg.",
            "api": "POST /api/inspections/{id}/signoff (role=physical)",
            "data": "Sandbox: physical sudah signed",
            "note": "Physical check fokus suhu kemasan (es / dingin) + integritas plastik / dus.",
        },
        {
            "fitur": "Distribusi — Aggregate Hari Ini",
            "path": "Sidebar > Produksi & Distribusi > Distribusi (🚐) > [tab default] Aggregate Hari Ini",
            "url": "/distributions (tab=today)",
            "cara": (
                "1. Klik 'Distribusi'\n"
                "2. Tab default 'Aggregate Hari Ini'\n"
                "3. Tabel: Sekolah | Wave | Target Porsi | Tray Delivered | Confirmed | Status\n"
                "4. Footer summary: Total target | Total confirmed | Confirm rate %"
            ),
            "hasil": "Real-time monitoring distribusi hari ini. Refresh auto tiap 30s.",
            "api": "GET /api/distributions/today",
            "data": "Sandbox: 3 sekolah, target porsi sesuai student_count",
            "note": "Wave classifier auto. Override manual butuh code change.",
        },
        {
            "fitur": "Distribusi — Wave 1 & 2",
            "path": "Distribusi > [tab] Wave 1 & 2",
            "url": "/distributions (tab=waves)",
            "cara": (
                "1. Tab 'Wave 1 & 2'\n"
                "2. Section Wave 1 (08:00 dispatch): list sekolah PAUD/TK/SD-low\n"
                "3. Section Wave 2 (10:00 dispatch): list sekolah SD-high/SMP/SMA\n"
                "4. Per sekolah: tombol 'Dispatch' (kalau status pending)\n"
                "5. Klik dispatch → status='dispatched', timer mulai untuk countdown delivery"
            ),
            "hasil": "Status delivery → 'dispatched'. Notif ke guru sekolah (kalau channel WA/SMS configured).",
            "api": "POST /api/distributions/{id}/dispatch",
            "data": "Sandbox: 3 sekolah split — TK Mawar wave 1, SD Tunas wave 1, SMP Bina wave 2",
            "note": "Wave classification rule: age_group startswith 'PAUD'/'TK' atau 'SD (7-9' → wave 1.",
        },
        {
            "fitur": "Distribusi — Sisa Porsi (Leftover)",
            "path": "Distribusi > [tab] Sisa Porsi",
            "url": "/distributions (tab=leftovers)",
            "cara": (
                "1. Tab 'Sisa Porsi'\n"
                "2. Tombol 'Tambah Leftover' → form:\n"
                "   - 'Date' (default today)\n"
                "   - 'School' (optional)\n"
                "   - 'Porsi Leftover' (jumlah)\n"
                "   - 'Reason' dropdown: absent / over-portion / quality / other\n"
                "   - 'Notes'\n"
                "3. Save"
            ),
            "hasil": "Waste log; weekly waste report aggregate. Mempengaruhi compliance score.",
            "api": "POST /api/distributions/leftovers, GET /leftovers",
            "data": "Sandbox: 4 porsi leftover per hari, reason=absent",
            "note": "Reason 'absent' = siswa tidak masuk; 'over-portion' = porsi >student_count.",
        },
        {
            "fitur": "ASLAP — Checklist Hari Ini",
            "path": "Sidebar > Lapangan & Monitoring > ASLAP Daily (📝) > [tab default] Checklist Hari Ini",
            "url": "/aslap (tab=checklist)",
            "cara": (
                "1. Klik 'ASLAP Daily'\n"
                "2. Tab default 'Checklist Hari Ini'\n"
                "3. List 7 item daily ops template (kebersihan dapur, suhu kulkas, APD, sanitasi tangan, dll)\n"
                "4. Per item: checkbox + textarea notes (optional)\n"
                "5. Tombol 'Submit' di footer → save semua\n"
                "6. Date picker untuk lihat checklist hari lain (history read-only)"
            ),
            "hasil": (
                "• Row aslap_checklist_results INSERT per item (target_date, key, checked, notes)\n"
                "• Aggregate masuk weekly report\n"
                "• Score compliance: % checked vs total"
            ),
            "api": "GET /api/aslap/checklists/today, POST /submit",
            "data": "Sandbox: 7 items all checked = score 100%",
            "note": "Template checklist managed di backend (default seed). Custom template butuh head_sppg + endpoint /aslap/template.",
        },
        {
            "fitur": "ASLAP — Water Quality",
            "path": "Sidebar > ASLAP Daily > [tab] Water Quality",
            "url": "/aslap (tab=water)",
            "cara": (
                "1. Tab 'Water Quality'\n"
                "2. Tombol 'Tambah Reading' → form:\n"
                "   - 'Source' dropdown: kran_dapur / galon / sumur / pdam\n"
                "   - 'pH' (number, range 0-14)\n"
                "   - 'TDS' (ppm, integer)\n"
                "   - 'Notes'\n"
                "3. Save\n"
                "4. Tabel history dengan warning kalau pH <6.5 / >8.5 atau TDS >500"
            ),
            "hasil": "Water log harian. Compliance check: pH 6.5-8.5, TDS <500 ppm (Permenkes).",
            "api": "POST /api/aslap/water-quality, GET /water-quality",
            "data": "Sandbox: pH 7.2 / TDS 180 (compliant)",
            "note": "Out-of-spec auto-trigger notif severity high ke head_sppg.",
        },
        {
            "fitur": "ASLAP — Production Obs",
            "path": "Sidebar > ASLAP Daily > [tab] Production Obs",
            "url": "/aslap (tab=obs)",
            "cara": (
                "1. Tab 'Production Obs'\n"
                "2. Tombol 'Tambah Observasi' → form:\n"
                "   - 'Category': kebersihan / suhu / SOP / lain-lain\n"
                "   - 'Severity': low / medium / high\n"
                "   - 'Description' (free text)\n"
                "3. Save\n"
                "4. Tabel history grouped by date"
            ),
            "hasil": "Observation log harian. Severity high notify head_sppg langsung.",
            "api": "POST /api/aslap/observations, GET /observations",
            "data": "Sandbox: 1 obs per hari ('dapur bersih, suhu OK', severity=low)",
            "note": "Field observation, bukan formal audit. Untuk audit pakai inspection.",
        },
        {
            "fitur": "ASLAP — Komunikasi Sekolah",
            "path": "Sidebar > ASLAP Daily > [tab] Komunikasi Sekolah",
            "url": "/aslap (tab=comm)",
            "cara": (
                "1. Tab 'Komunikasi Sekolah'\n"
                "2. Tombol 'Log Komunikasi' → form:\n"
                "   - 'School' dropdown\n"
                "   - 'Channel': wa / telp / visit / email\n"
                "   - 'Summary' (ex: 'Konfirmasi jadwal kirim besok 10:00')\n"
                "3. Save\n"
                "4. Tabel history per sekolah"
            ),
            "hasil": "Log percakapan dengan guru sekolah. Audit trail untuk koordinasi distribusi.",
            "api": "POST /api/aslap/comm-logs, GET /comm-logs",
            "data": "Sandbox: 1 log SD Tunas Bangsa via WA per hari",
            "note": "Channel 'visit' implies kunjungan fisik; auto-log GPS kalau enable (tidak default).",
        },
        {
            "fitur": "ASLAP — Generate Weekly Report",
            "path": "Sidebar > ASLAP Daily > [tab] Weekly Report > [tombol] 'Generate Laporan'",
            "url": "/aslap (tab=reports)",
            "cara": (
                "1. Tab 'Weekly Report'\n"
                "2. Date picker: pilih week (Mon-Sun)\n"
                "3. Tombol 'Generate Laporan' → backend aggregate:\n"
                "   - Checklist compliance %\n"
                "   - Water quality summary\n"
                "   - Observations summary\n"
                "   - Comm logs count\n"
                "   - Leftovers total\n"
                "4. Preview muncul di modal\n"
                "5. Tombol 'Submit' → kirim ke head_sppg untuk sign-off"
            ),
            "hasil": (
                "• aslap_reports row INSERT, status='draft' → 'pending_signoff'\n"
                "• Notif ke head_sppg\n"
                "• Setelah signed, masuk Compliance Bundle"
            ),
            "api": "POST /api/aslap/reports/generate, /submit",
            "data": "Sandbox: belum generate (perlu manual)",
            "note": "Report PDF-ready setelah signed. Export via Compliance Bundle.",
        },
        {
            "fitur": "Capture Permintaan Khusus Siswa",
            "path": "Sidebar > Menu & Gizi > Approval Menu > [section] Permintaan Khusus > [tombol] 'Tambah'",
            "url": "/menu-approval",
            "cara": (
                "1. Buka 'Approval Menu'\n"
                "2. Section 'Permintaan Khusus Siswa'\n"
                "3. Tombol 'Tambah' → form:\n"
                "   - 'School' dropdown\n"
                "   - 'Kelas' (text, ex: 'TK A')\n"
                "   - 'Student Name'\n"
                "   - 'Request Text' (ex: 'Anak alergi telur')\n"
                "4. Save"
            ),
            "hasil": "Request status='open'. Notif ke head_sppg / Ahli Gizi untuk resolve.",
            "api": "POST /api/student-requests",
            "data": "Sandbox: 1 request 'alergi telur' per hari",
            "note": "Capture: aslap atau Ahli Gizi. Resolve: head_sppg atau Ahli Gizi.",
        },
    ]),

    # ═════════════════════════════════════════════════════════════════════
    ("🍳 HEAD_KITCHEN / Kepala Chef", [
        {
            "fitur": "Mulai Batch Produksi (Dry-Run + Real)",
            "path": "Sidebar > Produksi & Distribusi > Production (🍳) > [section] Menu Approved Hari Ini > tombol 'Start'",
            "url": "/production",
            "cara": (
                "1. Klik 'Production' di sidebar\n"
                "2. Section 'Menu Approved Hari Ini' — list menu dengan target_date hari ini\n"
                "3. Per menu: nama, cost/porsi, AKG, tombol 'Start Batch'\n"
                "4. Modal:\n"
                "   - 'Target Porsi' (default = sum student_count semua sekolah)\n"
                "   - Toggle 'Dry Run' → preview FIFO plan tanpa execute\n"
                "5. Klik 'Confirm' → batch terbuat status='started'\n"
                "6. Backend FIFO debit stock items by storage_routing"
            ),
            "hasil": (
                "• Row production_batches INSERT (status=started, started_at=now, sop_max_minutes)\n"
                "• Items debit FIFO (by created_date_receiving)\n"
                "• Notif ke Ahli Gizi: 'Batch ready, perlu QC nanti'\n"
                "• Timer 4-6 jam SOP BGN mulai"
            ),
            "api": "POST /api/production/batches (body: menu_plan_id, target_porsi, dry_run)",
            "data": "Sandbox: batch id=2 sudah ended (5 porsi)",
            "note": "Dry-run aman; tidak debit stock. Real start lock items selama batch active.",
        },
        {
            "fitur": "Tablet Scan — Step Processing",
            "path": "Sidebar > Produksi > Production > [section bawah] Tablet Scan",
            "url": "/production (section 'Tablet Scan')",
            "cara": (
                "1. Section 'Tablet Scan — Step Processing (JWT auth)' di bagian bawah halaman\n"
                "2. Tombol 'Start' → kamera tablet aktif (request permission)\n"
                "3. Scan QR 'BHN-XXXXXXXX' di container\n"
                "4. Pilih step (radio): Receiving / Processing / Packing / Delivery\n"
                "5. Submit → backend update item status sesuai step\n"
                "6. Tombol 'Stop' untuk matikan kamera"
            ),
            "hasil": (
                "• scans table INSERT (kitchen_id, code, step, label, created_at)\n"
                "• items table column update (processing/packing/delivery=true)\n"
                "• Sync ke local_scans.db (offline-tolerant)"
            ),
            "api": "POST /api/scans (JWT required for Processing step)",
            "data": "Sandbox: belum ada scan (perlu QR fisik)",
            "note": "Step Receiving / Packing / Delivery juga bisa via Android scanner standalone (pakai scanner_key kitchen).",
        },
        {
            "fitur": "Akhiri Batch",
            "path": "Production > [section] Batch Aktif > batch active > tombol 'End'",
            "url": "/production",
            "cara": "1. Section 'Batch Aktif' di tengah halaman.\n2. Klik tombol 'End' (merah) di baris batch qc_passed.\n3. Confirm.",
            "hasil": (
                "• Status batch → 'ended', ended_at=now\n"
                "• Stock konsumsi finalized\n"
                "• Notif ke ASLAP: 'Batch ready, mulai distribusi'"
            ),
            "api": "POST /api/production/batches/{id}/end",
            "data": "Sandbox: batch id=2 sudah ended",
            "note": "End hanya bisa kalau batch sudah qc_passed. End batch started → error 400.",
        },
        {
            "fitur": "Receiving (Quick Add) — buat item ad-hoc",
            "path": "Sidebar > Penerimaan Bahan > Receiving (Quick) (↓)",
            "url": "/receiving",
            "cara": (
                "1. Klik 'Receiving (Quick)' di sidebar\n"
                "2. Form quick add bahan tanpa PO:\n"
                "   - 'Item Name'\n"
                "   - 'Weight Grams'\n"
                "   - 'Unit'\n"
                "   - 'Storage Routing' (cook_immediate/refrigerate/freeze)\n"
                "3. Save → container ID auto-generate, label print"
            ),
            "hasil": "Item ad-hoc masuk stock (untuk emergency / manual receipt).",
            "api": "POST /api/items",
            "data": "Sandbox: belum ada item ad-hoc",
            "note": "Bypass workflow inspection — cuma buat emergency. Audit log INSERT.",
        },
        {
            "fitur": "Edit Items / Edit Foods Override",
            "path": "(via API) atau menu master tertentu kalau UI tersedia",
            "url": "/items (kalau perm items.edit)",
            "cara": "CRUD master bahan: name, weight_grams, unit, default_supplier_id.",
            "hasil": "Master items updated.",
            "api": "PUT /api/items/{id}",
            "data": "Sandbox: items dibuat by inspection accept",
            "note": "Foods nutrition override (kalori per 100g) butuh perm foods.edit (Ahli Gizi).",
        },
    ]),

    # ═════════════════════════════════════════════════════════════════════
    ("🏫 PUBLIC / Guru Sekolah (TANPA LOGIN)", [
        {
            "fitur": "Konfirmasi Terima Makanan",
            "path": "Public link (TANPA sidebar/login)",
            "url": "/countdown/<tray_id>",
            "cara": (
                "1. Buka link 'http://localhost:5173/countdown/<tray_id>' (dari QR / SMS / WA)\n"
                "2. Halaman publik tanpa header/sidebar\n"
                "3. Form:\n"
                "   - 'Nama Guru' (free text)\n"
                "   - 'Confirmed Count' (jumlah ompreng diterima)\n"
                "   - 'Notes' (optional, ex: '12 ompreng, 3 siswa absent')\n"
                "4. Tombol 'Konfirmasi'\n"
                "5. Sukses screen muncul dengan ID konfirmasi"
            ),
            "hasil": (
                "• delivery_confirmations row INSERT (no auth, public endpoint)\n"
                "• Tray status → 'received' kalau confirmed_count >= target\n"
                "• Aggregate dashboard /distributions update real-time via SSE\n"
                "• Notif ke ASLAP + Akuntan: '<School> konfirmasi N porsi'"
            ),
            "api": "POST /api/countdown/{tray_id}/confirm-receipt (NO AUTH)",
            "data": "Sandbox: belum ada tray confirmation (perlu deploy menu/batch ke distribusi)",
            "note": "Endpoint publik—tidak ada auth required. Di-secure pakai tray_id (UUID-like) yang random.",
        },
    ]),
]


# ── Excel writer ───────────────────────────────────────────────────────────
def render_sheet(ws: Worksheet, group_title: str, rows: list[dict],
                 row_offset: int) -> int:
    """Render a role group: title bar + header row + data rows. Returns next row idx."""
    # Title bar (merged)
    ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=len(COLS))
    cell = ws.cell(row=row_offset, column=1, value=group_title)
    cell.fill = GROUP_FILL
    cell.font = GROUP_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_offset].height = 24
    row_offset += 1

    # Header row
    for ci, (name, _w) in enumerate(COLS, start=1):
        cell = ws.cell(row=row_offset, column=ci, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row_offset].height = 28
    row_offset += 1

    # Data rows
    for i, r in enumerate(rows, start=1):
        is_alt = (i % 2 == 0)
        values = [
            i,
            r["fitur"],
            r["path"],
            r["url"],
            r["cara"],
            r["hasil"],
            r["api"],
            r["data"],
            r["note"],
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=row_offset, column=ci, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
            if is_alt:
                cell.fill = ALT_FILL
        # Compute row height: estimate based on longest cell wrapped
        max_lines = max(
            (len(str(v).split("\n")) +
             sum(len(line) // 60 for line in str(v).split("\n")))
            for v in values
        )
        ws.row_dimensions[row_offset].height = max(18, min(220, max_lines * 14))
        row_offset += 1

    # Spacer row
    row_offset += 1
    return row_offset


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Fitur DPMBG"

    # Set column widths
    for ci, (_n, width) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Top header / metadata
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    title = ws.cell(row=1, column=1, value="DPMBG — Catalog Fitur (Group by Role) · Test Sandbox")
    title.font = Font(bold=True, size=14, color="1E3A8A")
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
    sub = ws.cell(row=2, column=1, value=(
        "Login: http://localhost:5173/  |  testadmin / testadmin123 / org_slug=testsandbox  "
        "|  Path = sidebar/page/tab persis seperti di UI"
    ))
    sub.font = Font(italic=True, size=10, color="475569")
    sub.alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[2].height = 18

    # Freeze header
    row = 4
    for title_, items in ROLE_GROUPS:
        row = render_sheet(ws, title_, items, row)

    ws.freeze_panes = "C5"

    out_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "docs", "DPMBG_Features.xlsx",
    )
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    p = build()
    print(f"[OK] Excel saved: {p}")
